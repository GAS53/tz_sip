from collections import defaultdict
import os

from pydub import AudioSegment
from rest_framework import generics
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q
from django.core.files.storage import default_storage
from rest_framework.views import APIView
from rest_framework.parsers import FileUploadParser
from openpyxl import load_workbook
from rest_framework_simplejwt.models import TokenUser

from config import settings
from baseapp.models import CustomUser, Project, Call
from callapp.serializers import ProjectSerialiser, RunSerializer


class CreateProjectAPI(generics.CreateAPIView):
    permission_classes = (IsAuthenticated,)
    serializer_class = ProjectSerialiser

    def post(self, request, *args, **kwargs):
        from_token = TokenUser(token=request.auth)
        user = CustomUser.objects.filter(id=from_token.id).first()
        data = self.new_project(request)
        proj = Project.objects.create(user_id=user.id, title=data['title'], description=data['description'],count_cols=data['count_cols'])
        return Response({'new_project': proj.title})

    def new_project(self, request):
        serialzer = self.serializer_class(data=request.data)
        serialzer.is_valid(raise_exception=True)
        data = defaultdict(str)
        data.update(serialzer.data)
        data['count_cols'] = 1
        return data
        

class GetAllProjects(generics.ListAPIView):
    permission_classes = (IsAuthenticated,)
    serializer_class = ProjectSerialiser

    def get_queryset(self):
        queryset = Project.objects.all()
        if not self.request.user.is_staff:
            if queryset is None:
                return Response({'error': 'у пользователя нет проектов'})
            queryset = queryset.filter(user_id=self.request.user.id)
        return queryset

    def get(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)
        


def check_project(request, pk):
    if not pk:
        return Response({'error': 'Не определен pk проекта'})
    user = CustomUser.objects.filter(id=request.user.id).first()
    project = Project.objects.filter(id=pk).first()

    if not user:
        return Response({'error': "ошибка задания пользователя"})
    if not project:
        return Response({'error': "ошибка поиска проекта"})
    if project.user.id == user.id or user.is_staff:
        return project
    return Response({'error': "у пользователя нет такого проекта"})
    


class CheckProjectAPI(generics.RetrieveAPIView):
    permission_classes = (IsAuthenticated,)
    
    def get(self, request, *args, **kwargs):
        pk = kwargs.get('pk', None)
        res = check_project(request, pk)
        if isinstance(res, Response):
            return res

        q_all = Q(project_id=pk)
        q_is_call = Q(is_call=True)
        q_is_dial = Q(is_dial=True)
                
        all_numbers = Call.objects.filter(q_all).count()
        all_is_calling = Call.objects.filter(q_all & q_is_call).count()
        all_is_dial = Call.objects.filter(q_all & q_is_dial).count()

        if all_numbers == 0:
            return Response({'error': 'нет номеров в проекте'})

        res ={}
        res['all_numbers'] = all_numbers
        res['all_is_calling'] = all_is_calling # без учета повторных звонков для недозвонившихся
        res['all_is_dial'] = all_is_dial

        return Response({'res': res})


def get_project_from_file(val):
    try:
        val = int(val.split('_')[0])
        if val is None:
            return Response({'error': 'в начале файла не указан id проекта'})
        return val
    except Exception as e:
        return Response({'error': 'ошибка в названии файла'})
    

def rename_wav_to_mp3(file):
    file_name = os.path.splitext(file)[0]
    return f'{file_name}.wav'


def sound_converter(file):
    print(file)
    sound = AudioSegment.from_mp3(file)
    # sound.sample_width = 1
    # sound.frame_rate = 8000
    sound.set_channels = 1
    new_file = rename_wav_to_mp3(file)
    sound.export(new_file, format="wav")
    os.remove(file)


class UploadFileView(APIView):
    parser_class = (FileUploadParser,)
    permission_classes = (IsAuthenticated,)

    def post(self, request, *args, **kwargs):
        file = request.data['file']
        name = str(file)
        id_project = get_project_from_file(name)
        project = Project.objects.filter(id=id_project).first()
        if not project:
            return Response({'error': f'проекта с номером {id_project} нет'})
        if project.user.id != request.user.id:
            return Response({'error': f'вы пытаетесь добавить файл не в свой проект'})
        new_name = rename_wav_to_mp3(file.name)
        project.file = new_name
        project.save()
        default_storage.save(file.name, file)
        path = os.path.join(os.getcwd(), settings.MEDIA_ROOT, file.name)
        sound_converter(path)
        return Response({'res': 'файл сохранен'})




def dict_creater(project_id, phone, count_cols=None):
    di = {}
    di['phone'] = phone
    di['project_id'] = project_id
    if count_cols:
        di['count_cols'] = count_cols
    return di


def pars_row(ws, project_id):
    for row in ws.iter_rows():
        if row[1].value:
            yield Call(project_id=project_id, phone=row[0].value, count_cols=row[1].value)
        else: 
            yield Call(project_id=project_id, phone=row[0].value)

def unpack_to_base(path, project_id):
    print(f' паршу файл {path}')
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    region_list = []
    for row in pars_row(ws, project_id):
        region_list.append(row)
    unpack_calls(region_list)
    return region_list


def unpack_calls(calls):
    try:
        Call.objects.bulk_create(calls)
    except:
        return Response({'error': f'файл имеет неправильное форматирование'})


class AddPhonesAPI(APIView):
    parser_class = (FileUploadParser,)
    permission_classes = (IsAuthenticated,)

    def post(self, request, *args, **kwargs):
        file = request.data['file']
        if not file.name.endswith('xlsx'):
            return Response({'error': 'не подходит формат файла - должен быть xlsx'})
        id_project = get_project_from_file(file.name)
        project = Project.objects.filter(id=id_project).first()
        if not project:
            return Response({'error': f'проекта с номером {id_project} нет'})
        if project.user.id != request.user.id:
            return Response({'error': f'вы пытаетесь добавить файл не в свой проект'})
        project.file = file.name
        project.save()
        default_storage.save(file.name, file)
        path = os.path.join(os.getcwd(), settings.MEDIA_ROOT, file.name)
        try:
            unpack_to_base(path, project.id)
        finally:
            os.remove(path)
        return Response({'res': 'файл добавлен в базу'})
    

from django.db.models import F

class RunCallingAPI(generics.RetrieveAPIView):
    permission_classes = (IsAuthenticated,)
    serializer_class = ProjectSerialiser

    def get(self, request, *args, **kwargs):
        pk = kwargs.get('pk', None)
        project = check_project(request, pk)
        if isinstance(project, Response):
            return project
        activ_project_calls = Call.objects.filter(project_id=project.id, num_call_now__lt=F('count_cols'))
        numbers = [number.phone for number in activ_project_calls]

        return Response({'call numbers': numbers})


